import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

print("Creating properly formatted Excel file...")

# Load the CSV
df = pd.read_csv('master_indian_weather_dataset.csv')

# Save to Excel
excel_file = 'master_indian_weather_dataset.xlsx'
df.to_excel(excel_file, index=False, sheet_name='Weather Data')

# Load the workbook to format it
wb = load_workbook(excel_file)
ws = wb['Weather Data']

# Set column widths
column_widths = {
    'A': 12,  # date
    'B': 15,  # location
    'C': 10,  # state
    'D': 16,  # max_temperature
    'E': 16,  # min_temperature
    'F': 16,  # avg_temperature
    'G': 12,  # temp_range
    'H': 10,  # humidity
    'I': 10,  # rainfall
    'J': 12,  # wind_speed
    'K': 12,  # risk_level
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Make header row bold
header_font = Font(bold=True, size=11)
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Save the formatted workbook
wb.save(excel_file)

print(f"✅ Excel file created: {excel_file}")
print(f"   - All columns properly sized")
print(f"   - Header row formatted")
print(f"   - Total records: {len(df):,}")
print(f"\nOpen '{excel_file}' instead of the CSV file!")
